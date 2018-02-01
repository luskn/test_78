# -*- coding: utf-8 -*-

import openpyxl as xl
from os import path

def check_good_date(in_good, in_rev_date):
    """For one good, can be only one revenue by date"""
    from revenue.models import Revenue

    try:
        Revenue.objects.get(good=in_good, date_sold=in_rev_date)
    except Revenue.DoesNotExist:
        return False
    else:
        return True

def load_from_xlsx_to_revenue(fpath, fname):
    """Load data from xlsx file to model revenue (db table revenue)"""
    from revenue.models import Revenue

    wb = xl.load_workbook(path.join(fpath, fname))
    ws = wb.active

    dates_list = ws[1][1:]
    for row in ws.iter_rows(min_row=2):
        good = row[0].value
        if not good:  # Если закончились товары, то прерываем обход
            break
        for ind, cell in enumerate(row[1:]):
            isExist = check_good_date(in_good=good, in_rev_date=dates_list[ind].value)
            if not isExist:
                print(good, cell.value, dates_list[ind].value)
                rev = Revenue(good=good,
                              revenue=cell.value,
                              date_sold=dates_list[ind].value)
                rev.save()


if __name__ == "__main__":
    import django, os

    os.environ.setdefault("DJANGO_SETTINGS_MODULE", 'test_78.settings')
    django.setup()
    load_from_xlsx_to_revenue(r"C:\Users\luskn\Downloads\Telegram Desktop\backend_test_78", "Здадание 1.xlsx")


